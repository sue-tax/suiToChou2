'''
Created on 2021/02/09
出納帳
逢ふことの絶えてしなくはなかなかに
　　　　　　　人をも身をも恨みざらまし
　　　　　　　　　　　中納言朝忠
@author: sue-t
'''
# from pymupdf._mupdf import delete_pdf_hint_page
# from SuiToChou_v006 import F_HIZUKE_COLUMN, F_KARIKATA_KAMOKU_COLUMN, F_KARIKATA_HOJO_KAMOKU_COLUMN,\
#     F_KARIKATA_KINGAKU_COLUMN
# from pickle import NONE

'''
出納帳のExcelデータを元に、
残高試算表・仕訳帳・総勘定元帳・補助元帳のExcelファイルと
翌期入力用の出納帳のExcelファイルを
作成する

dataframe にデータを放り込む
現金・預金間で重複するデータは削除する
伝票番号を自動生成する
複合仕訳と現金・預金間で重複データを削除する
複合仕訳の伝票番号はマイナス　（上手な対応方法を考える）

部門、税区分、税額は対応しない
'''

'''
TODO
ドロップダウンリストの設定確認
エラーチェック
　日付の範囲
　設定シートの合計額
収益、費用への対応
設定シートの科目の順番自由化
各合計（販管費合計など）に対応できないか？
 +++XXXX+++を集計項目として活用する？
'''


import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl.reader
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta

import c
import d
import e

SETTEI_FILE_NAME = r'設定.xlsx'
SETTEI_SHEET_NAME = r'設定'
TANITSU_SHEET_NAME = r'単一仕訳'
FUKUGOU_SHEET_NAME = r'複合仕訳'

KIHON_SHEET_NAME = '基本'

# 出納帳シート
HIZUKE = '日付'
DENPYOU_BANGOU = '番号'   # '伝票番号'
AITE_KAMOKU = '相手科目'
AITE_HOJO_KAMOKU = '相手補助'   # '相手補助科目'
TEKIYOU = '摘要'
TEKIYOU2 = '摘要２'
NYUKIN = '入金'
SHUKKIN = '出金'
ZANDAKA = '残高'

S_KAMOKU_HEADER_ROW = 0
S_KISHU_ZANDAKA_ROW = 2

S_HIZUKE_COLUMN = 0         # ExcelからDataFrameの読込み　最初は０
S_AITE_KAMOKU_COLUMN = 1
S_AITE_HOJO_KAMOKU_COLUMN = 2
S_TEKIYOU_COLUMN = 3
S_TEKIYOU2_COLUMN = 4
S_NYUKIN_COLUMN = 5
S_SHUKKIN_COLUMN = 6
S_ZANDAKA_COLUMN = 7

# 振替伝票シート
KARIKATA_KAMOKU = '借方科目'
KARIKATA_HOJO_KAMOKU = '借方補助'   # '借方補助科目'
KARIKATA_KINGAKU = '借方金額'
KASHIKATA_KAMOKU = '貸方科目'
KASHIKATA_HOJO_KAMOKU = '貸方補助'    # '貸方補助科目'
KASHIKATA_KINGAKU = '貸方金額'

F_TITLE_ROW = 0

F_HIZUKE_COLUMN = 0
F_KARIKATA_KAMOKU_COLUMN = 1
F_KARIKATA_HOJO_KAMOKU_COLUMN = 2
F_KARIKATA_KINGAKU_COLUMN = 3
F_KASHIKATA_KAMOKU_COLUMN = 4
F_KASHIKATA_HOJO_KAMOKU_COLUMN = 5
F_KASHIKATA_KINGAKU_COLUMN = 6
F_TEKIYOU_COLUMN = 7
F_TEKIYOU2_COLUMN = 8

HOJO_KAMOKU = '補助科目'

# 基本シート
DANTAI_MEI_ROW = 1
KISHU_BI_ROW = 2
KIMATSU_BI_ROW = 3
KAMOKU_TITLE_ROW = 4

KAMOKU_COLUMN = 1       # Excelシートの読込み　最初は１
HOJO_KAMOKU_COLUMN = 2
KAMOKU_ZANDAKA_COLUMN = 3
HOJO_KAMOKU_ZANDAKA_COLUMN = 4
SUITOU_COLUMN = 5
TAISHAKU_KUBUN_COLUMN = 6

DANTAIMEI = '団体名'
HOUJINMEI = '法人名'
KISHU_BI = '期首日'
KIMATSU_BI = '期末日'
KISHU_ZANDAKA = '期首残高'
HOJO_KISHU_ZANDAKA = '補助残高'
SUITOU = '出納'
KUBUN = '区分'

KAMOKU_HOJO_GOUKEI = '【合計】'

SUITOU_ARI = '有'

TAISHAKU_KUBUN_KARI = '借'
TAISHAKU_KUBUN_KASHI = '貸'

TAISHAKU_KUBUN_SHISAN = '資産'
TAISHAKU_KUBUN_FUSAI = '負債'
TAISHAKU_KUBUN_SHIHON = '資本'
TAISHAKU_KUBUN_JUNSHISAN = '純資産'
TAISHAKU_KUBUN_SHUNYU = '収入'
TAISHAKU_KUBUN_SHISHUTSU = '支出'
TAISHAKU_KUBUN_SHUSHI = '収支'
TAISHAKU_KUBUN_SHUEKI = '収益'
TAISHAKU_KUBUN_HIYOU = '費用'
TAISHAKU_KUBUN_RIEKI = '利益'

TANI_EN = '(単位:円)'

SHOKUCHI = '諸口'         # 複合仕訳の相手科目名
FUKUGOU_GOUKEI = '合計'   # 複合仕訳の終わりを示す


global kishu_bi, kimatsu_bi


def read_suitou(excel_file_name, sheet_name,
            kamoku_mei, hojo_kamoku_mei):
    '''
    Excelファイル内の出納帳シートを読込み、
    仕訳データを作成する。

    Parameters
    ----------
    excel_file_name : str
        読込むExcelファイル名。
    sheet_name : str
        出納帳データが入力されているシート名。
    kamoku_mei : str
        出納帳の科目名。
    hojo_kamoku_mei : str
        出納帳の補助科目名。

    Returns
    df_shwake : DataFrame
        仕訳データ。
    -------
    '''
    d.dprint_method_start()

    openpyxl.reader.excel.warnings.simplefilter('ignore') # warning 対策　入力規則無視
    if not os.path.isfile(excel_file_name):
        e.eprint('ファイルがありません', excel_file_name)
        exit()
    try:
        df_suitou = pd.read_excel(excel_file_name, sheet_name, \
                header=0,
                skiprows=[S_KAMOKU_HEADER_ROW, S_KISHU_ZANDAKA_ROW],
                usecols=[S_HIZUKE_COLUMN,
                        S_AITE_KAMOKU_COLUMN, S_AITE_HOJO_KAMOKU_COLUMN,
                        S_TEKIYOU_COLUMN, S_TEKIYOU2_COLUMN,
                        S_NYUKIN_COLUMN, S_SHUKKIN_COLUMN],
                engine='openpyxl')
    except:
        msg = "ファイル「{}」にシート「{}」が必要です。" \
                .format(excel_file_name, sheet_name)
        e.eprint('シートがありません', msg)
        exit()

    # 日付、相手科目の列を見て、空行と判断し削除する
    df_suitou.dropna(subset=[HIZUKE, AITE_KAMOKU],
            how='all', inplace=True)    # error
    # 要検討
    df_suitou.insert(1, DENPYOU_BANGOU, df_suitou.index+1)
    # 空欄にデータを補充
    df_suitou.fillna({ \
            AITE_HOJO_KAMOKU: '',
            TEKIYOU: '',
            TEKIYOU2: '',
            NYUKIN: 0,
            SHUKKIN: 0},
            inplace=True)
    df_suitou[[NYUKIN, SHUKKIN]] \
            = df_suitou[[NYUKIN, SHUKKIN]].astype('int')
    df_suitou[[TEKIYOU, TEKIYOU2]] \
            = df_suitou[[TEKIYOU, TEKIYOU2]].astype('str')
    pd.to_datetime(df_suitou[HIZUKE], format="%Y-%m-%d")   # 20210823

    # 日付が期間内かをチェック
    str_query = '{} < "{}" or "{}" < {}' \
            .format(HIZUKE, kishu_bi, kimatsu_bi, HIZUKE)
    d.dprint(str_query)
    df_out = df_suitou.query(str_query)
    d.dprint(df_out)
    if len(df_out) > 0:
        for _index, row in df_out.iterrows():
            d.dprint(_index)
            msg = "シート{}の{}行目の日付{}は範囲外です。" \
                    .format(sheet_name,
                    row[DENPYOU_BANGOU]+3 , row[HIZUKE])
            e.eprint('データが間違っています', msg)
        exit(-1)
    # TODO 勘定科目、補助科目　チェック
    # 補助科目があれば、勘定科目・補助科目セットでチェック
    # 補助科目がなければ、勘定科目だけでチェック
    # read_kihon で作ったkamoku_list, hojo_kamoku_listを利用する
    
    d.dprint(df_suitou[AITE_HOJO_KAMOKU]) # "相手補助科目"])   # AITE_HOJO_KAMOKU])
    df_nyukin = df_suitou[df_suitou[NYUKIN] != 0]
    df_nyukin.insert(2, KARIKATA_KAMOKU, kamoku_mei)
    df_nyukin.insert(3, KARIKATA_HOJO_KAMOKU, hojo_kamoku_mei)
    df_nyukin.insert(4, KARIKATA_KINGAKU, df_nyukin[NYUKIN])
    df_nyukin_new = df_nyukin.rename(columns={
            AITE_KAMOKU: KASHIKATA_KAMOKU,
            AITE_HOJO_KAMOKU: KASHIKATA_HOJO_KAMOKU,
            NYUKIN: KASHIKATA_KINGAKU})
    del df_nyukin
    df_nyukin = df_nyukin_new
    df_nyukin[TEKIYOU] = df_nyukin[TEKIYOU] + ' ' + df_nyukin[TEKIYOU2]
    df_nyukin = df_nyukin.reindex([HIZUKE, DENPYOU_BANGOU,
            KARIKATA_KAMOKU, KARIKATA_HOJO_KAMOKU, KARIKATA_KINGAKU,
            KASHIKATA_KAMOKU, KASHIKATA_HOJO_KAMOKU, KASHIKATA_KINGAKU,
            TEKIYOU],
            axis='columns')

    df_shukkin = df_suitou[df_suitou[SHUKKIN] != 0]
    df_shukkin.insert(6, KASHIKATA_KAMOKU, kamoku_mei)
    df_shukkin.insert(7, KASHIKATA_HOJO_KAMOKU, hojo_kamoku_mei)
    df_shukkin.insert(8, KASHIKATA_KINGAKU, df_shukkin[SHUKKIN])
    df_shukkin_new = df_shukkin.rename(columns={
            AITE_KAMOKU: KARIKATA_KAMOKU,
            AITE_HOJO_KAMOKU: KARIKATA_HOJO_KAMOKU,
            SHUKKIN: KARIKATA_KINGAKU})
    del df_shukkin
    df_shukkin = df_shukkin_new
    df_shukkin[TEKIYOU] = df_shukkin[TEKIYOU] + ' ' + df_shukkin[TEKIYOU2]
    df_shukkin = df_shukkin.reindex([HIZUKE, DENPYOU_BANGOU,
            KARIKATA_KAMOKU, KARIKATA_HOJO_KAMOKU, KARIKATA_KINGAKU,
            KASHIKATA_KAMOKU, KASHIKATA_HOJO_KAMOKU, KASHIKATA_KINGAKU,
            TEKIYOU],
            axis='columns')

    df_shiwake = pd.concat([df_nyukin, df_shukkin])
    df_shiwake.sort_values(DENPYOU_BANGOU, inplace=True)

    del df_suitou
    del df_nyukin
    del df_shukkin
    d.dprint(df_shiwake)
    d.dprint_method_end()
    return df_shiwake

def read_tanitsu_shiwake(excel_file_name, sheet_name):
    '''
    Excelファイル内の単一仕訳シートを読込み、
    仕訳データを作成する。

    Parameters
    ----------
    excel_file_name : str
        読込むExcelファイル名。
    sheet_name : str
        単一仕訳データが入力されているシート名。

    Returns
    df_shwake : DataFrame
        仕訳データ。
    -------
    '''
    # d.dprint_method_start()
    openpyxl.reader.excel.warnings.simplefilter('ignore') # warning 対策　入力規則無視
    if not os.path.isfile(excel_file_name):
        e.eprint('ファイルがありません', excel_file_name)
        exit()
    try:
        df_furikae = pd.read_excel(excel_file_name, sheet_name, \
                header=0,
                skiprows=[F_TITLE_ROW],
                usecols=[F_HIZUKE_COLUMN,
                        F_KARIKATA_KAMOKU_COLUMN, F_KARIKATA_HOJO_KAMOKU_COLUMN,
                        F_KARIKATA_KINGAKU_COLUMN,
                        F_KASHIKATA_KAMOKU_COLUMN, F_KASHIKATA_HOJO_KAMOKU_COLUMN,
                        F_KASHIKATA_KINGAKU_COLUMN,
                        F_TEKIYOU_COLUMN, F_TEKIYOU2_COLUMN],
                engine='openpyxl')
    except:
        msg = "ファイル「{}」にシート「{}」が必要です。" \
                .format(excel_file_name, sheet_name)
        e.eprint('シートがありません', msg)
        exit()

    # 日付、借方金額の列を見て、空行と判断し削除する
    df_furikae.dropna(subset=[HIZUKE, KARIKATA_KINGAKU],
            how='all', inplace=True)
    # 要検討
    df_furikae.insert(1, DENPYOU_BANGOU, df_furikae.index+1)
    # 空欄にデータを補充
    df_furikae.fillna({ \
            KARIKATA_HOJO_KAMOKU: '',
            TEKIYOU: '',
            TEKIYOU2: '',
            KASHIKATA_HOJO_KAMOKU: ''},
            inplace=True)
    df_furikae[[KARIKATA_KINGAKU, KASHIKATA_KINGAKU]] \
            = df_furikae[[KARIKATA_KINGAKU, KASHIKATA_KINGAKU]] \
            .astype('int')
    pd.to_datetime(df_furikae[HIZUKE], format="%Y-%m-%d")   # 20210823

    if len(df_furikae) > 0:
        df_furikae[TEKIYOU] = df_furikae[TEKIYOU] + \
                ' ' + df_furikae[TEKIYOU2]
    df_furikae = df_furikae.reindex([HIZUKE, DENPYOU_BANGOU,
            KARIKATA_KAMOKU, KARIKATA_HOJO_KAMOKU, KARIKATA_KINGAKU,
            KASHIKATA_KAMOKU, KASHIKATA_HOJO_KAMOKU, KASHIKATA_KINGAKU,
            TEKIYOU],
            axis='columns')
    # d.dprint(df_furikae)
    # d.dprint_method_end()
    return df_furikae


def read_fukugou_shiwake(excel_file_name, sheet_name):
    '''
    Excelファイル内の振替仕訳シートを読込み、
    仕訳データを作成する。

    Parameters
    ----------
    excel_file_name : str
        読込むExcelファイル名。
    sheet_name : str
        振替仕訳データが入力されているシート名。

    Returns
    df_shwake : DataFrame
        仕訳データ。
    -------
    '''
    # d.dprint_method_start()
    openpyxl.reader.excel.warnings.simplefilter('ignore') # warning 対策　入力規則無視
    if not os.path.isfile(excel_file_name):
        e.eprint('ファイルがありません', excel_file_name)
        exit()
    try:
        df_fukugou = pd.read_excel(excel_file_name, sheet_name, \
                header=0,
                skiprows=[F_TITLE_ROW],
                usecols=[F_HIZUKE_COLUMN,
                        F_KARIKATA_KAMOKU_COLUMN, F_KARIKATA_HOJO_KAMOKU_COLUMN,
                        F_KARIKATA_KINGAKU_COLUMN,
                        F_KASHIKATA_KAMOKU_COLUMN, F_KASHIKATA_HOJO_KAMOKU_COLUMN,
                        F_KASHIKATA_KINGAKU_COLUMN,
                        F_TEKIYOU_COLUMN, F_TEKIYOU2_COLUMN],
                engine='openpyxl')
    except:
        msg = "ファイル「{}」にシート「{}」が必要です。" \
                .format(excel_file_name, sheet_name)
        e.eprint('シートがありません', msg)
        exit()

    df_furikae = pd.DataFrame(
            columns=[HIZUKE, DENPYOU_BANGOU,
                     KARIKATA_KAMOKU, KARIKATA_HOJO_KAMOKU, KARIKATA_KINGAKU,
                     KASHIKATA_KAMOKU, KASHIKATA_HOJO_KAMOKU, KASHIKATA_KINGAKU,
                     TEKIYOU, TEKIYOU2])
    # print(df_furikae)
    denpyou_bangou = -1
    data_list = []
    iter_fukugou = df_fukugou.itertuples()
    line_tuple = next(iter_fukugou)
    while not pd.isna(line_tuple[F_HIZUKE_COLUMN+1]):
        hizuke = line_tuple[F_HIZUKE_COLUMN+1]
        karikata_goukei = 0
        kashikata_goukei = 0
        # karikata_main = None
        # kashikata_main = None
        # data_list = []
        while line_tuple[F_HIZUKE_COLUMN+1] != FUKUGOU_GOUKEI:
            if not pd.isna(line_tuple[F_KARIKATA_KAMOKU_COLUMN+1]): 
                # if karikata_main == None:
                #     karikata_main = line_tuple[F_KARIKATA_KAMOKU_COLUMN+1]
                kingaku = line_tuple[F_KARIKATA_KINGAKU_COLUMN+1]
                karikata_goukei = karikata_goukei + kingaku
                new_data = { HIZUKE:hizuke,
                        DENPYOU_BANGOU:denpyou_bangou,
                        KARIKATA_KAMOKU: line_tuple[F_KARIKATA_KAMOKU_COLUMN+1],
                        KARIKATA_HOJO_KAMOKU: line_tuple[F_KARIKATA_HOJO_KAMOKU_COLUMN+1],
                        KARIKATA_KINGAKU: kingaku,
                        KASHIKATA_KAMOKU: SHOKUCHI,
                        KASHIKATA_HOJO_KAMOKU: '',
                        KASHIKATA_KINGAKU: kingaku,
                        TEKIYOU:  line_tuple[F_TEKIYOU_COLUMN+1],
                        TEKIYOU2:  line_tuple[F_TEKIYOU2_COLUMN+1]
                        }
                data_list.append(new_data)
            if not pd.isna(line_tuple[F_KASHIKATA_KAMOKU_COLUMN+1]): 
                # if kashikata_main == None:
                #     kashikata_main = line_tuple[F_KASHIKATA_KAMOKU_COLUMN+1]
                kingaku = line_tuple[F_KASHIKATA_KINGAKU_COLUMN+1]
                kashikata_goukei = kashikata_goukei + kingaku
                new_data = { HIZUKE:hizuke,
                        DENPYOU_BANGOU:denpyou_bangou,
                        KARIKATA_KAMOKU: SHOKUCHI,
                        KARIKATA_HOJO_KAMOKU: '',
                        KARIKATA_KINGAKU: kingaku,
                        KASHIKATA_KAMOKU: line_tuple[F_KASHIKATA_KAMOKU_COLUMN+1],
                        KASHIKATA_HOJO_KAMOKU: line_tuple[F_KASHIKATA_HOJO_KAMOKU_COLUMN+1],
                        KASHIKATA_KINGAKU: kingaku,
                        TEKIYOU:  line_tuple[F_TEKIYOU_COLUMN+1],
                        TEKIYOU2:  line_tuple[F_TEKIYOU2_COLUMN+1]
                        }
                data_list.append(new_data)
            try:
                line_tuple = next(iter_fukugou)
            except:
                # TODO 合計行がない
                msg = "{}シートの{}行目の次に合計行が必要です。" \
                        .format(sheet_name, line_tuple.Index)
                e.eprint('合計行がありません', msg)
                exit()
        if karikata_goukei != kashikata_goukei:
            msg = "{}シートの{}行目の合計が{}と{}です。" \
                    .format(sheet_name, line_tuple.Index,
                    karikata_goukei, kashikata_goukei)
            e.eprint('貸借金額が合っていません', msg)
            exit()
        if line_tuple[F_KARIKATA_KINGAKU_COLUMN+1] != karikata_goukei:
            msg = "{}シートの{}行目の合計{}と借方金額の集計額が{}です。" \
                    .format(sheet_name, line_tuple.Index,
                    karikata_goukei, line_tuple[F_KARIKATA_KINGAKU_COLUMN+1])
            e.eprint('合計金額が合っていません', msg)
            exit()
        if line_tuple[F_KASHIKATA_KINGAKU_COLUMN+1] != kashikata_goukei:
            msg = "{}シートの{}行目の合計{}と貸方金額の集計額が{}です。" \
                    .format(sheet_name, line_tuple.Index,
                    kashikata_goukei, line_tuple[F_KASHIKATA_KINGAKU_COLUMN+1])
            e.eprint('合計金額が合っていません', msg)
            exit()
        # df_furikae = pd.concat([df_furikae, pd.DataFrame(data_list)])
        try:        
            line_tuple = next(iter_fukugou)
        except:
            break
        denpyou_bangou -= 1
    
    df_furikae = pd.concat([df_furikae, pd.DataFrame(data_list)])
    # 空欄にデータを補充
    df_furikae.fillna({ \
            KARIKATA_HOJO_KAMOKU: '',
            TEKIYOU: '',
            TEKIYOU2: '',
            KASHIKATA_HOJO_KAMOKU: ''},
            inplace=True)
    df_furikae[[KARIKATA_KINGAKU, KASHIKATA_KINGAKU]] \
            = df_furikae[[KARIKATA_KINGAKU, KASHIKATA_KINGAKU]] \
            .astype('int')
    pd.to_datetime(df_furikae[HIZUKE], format="%Y-%m-%d")   # 20210823
    
    if len(df_furikae) > 0:
        df_furikae[TEKIYOU] = df_furikae[TEKIYOU] + \
                ' ' + df_furikae[TEKIYOU2]
    df_furikae = df_furikae.reindex([HIZUKE, DENPYOU_BANGOU,
            KARIKATA_KAMOKU, KARIKATA_HOJO_KAMOKU, KARIKATA_KINGAKU,
            KASHIKATA_KAMOKU, KASHIKATA_HOJO_KAMOKU, KASHIKATA_KINGAKU,
            TEKIYOU],
            axis='columns')
    # d.dprint(df_furikae)
    # d.dprint_method_end()
    return df_furikae


def ketsugou_shiwake(list_df_shiwake, list_suitou_kamoku,
             df_fukugou_shiwake):
    '''
    複数の仕訳データを結合し、重複仕訳を削除する。
    複合仕訳も結合する（伝票番号の振り直し、重複削除はしない）

    Parameters
    ----------
    list_df_shiwake : list of DataFrame
        仕訳データのリスト。
    list_suitou_kamoku : list of tuple of str
        出納帳データの科目・補助科目のタプルのリスト
        ex. [('現金', ''), ('普通預金', '三菱ＵＦＪ')]
    df_fukugou_shiwake : DataFrame
        複合仕訳データのリスト

    Returns
    df_shiwake : DataFrame
        結合した仕訳データ。
    -------
    '''
    # TODO 伝票番号　付け直し？

    # d.dprint_method_start()
    # d.dprint(list_df_shiwake)
    df_ketsugou = pd.concat(list_df_shiwake, ignore_index=True)
    df_ketsugou.sort_values([HIZUKE, DENPYOU_BANGOU],
            inplace=True)
    # 重複を削除
    index_pair_list = []
    for _, row in df_ketsugou.iterrows():
        if \
                ((row[KARIKATA_KAMOKU], row[KARIKATA_HOJO_KAMOKU]) \
                        in list_suitou_kamoku) \
                and \
                ((row[KASHIKATA_KAMOKU], row[KASHIKATA_HOJO_KAMOKU]) \
                        in list_suitou_kamoku):
            # 借方も貸方も、出納帳の科目である場合
            # 日付、借方科目、貸方科目、金額が一致するものをピックアップ
            index_pair = df_ketsugou.index[ \
                    (df_ketsugou[HIZUKE] == row[HIZUKE]) \
                    & (df_ketsugou[KARIKATA_KAMOKU] == row[KARIKATA_KAMOKU]) \
                    & (df_ketsugou[KARIKATA_HOJO_KAMOKU] == row[KARIKATA_HOJO_KAMOKU]) \
                    & (df_ketsugou[KASHIKATA_KAMOKU] == row[KASHIKATA_KAMOKU]) \
                    & (df_ketsugou[KASHIKATA_HOJO_KAMOKU] == row[KASHIKATA_HOJO_KAMOKU]) \
                    & (df_ketsugou[KARIKATA_KINGAKU] == row[KARIKATA_KINGAKU]) \
                    & (df_ketsugou[KASHIKATA_KINGAKU] == row[KASHIKATA_KINGAKU])]
            if (len(index_pair) % 2 == 0):
                # 同じ金額を２回下すことがあり得る。
                for index in range(0, len(index_pair), 2):
                    index_pair_list.append( \
                            (index_pair[index],
                            index_pair[index+1]))
            else:
                str_msg = "重複データが奇数です。\n{} {:,d} {:,d}\n {} {} {} {}\n{}". \
                        format(row[HIZUKE].strftime('%Y-%m-%d'),
                        row[KARIKATA_KINGAKU], row[KASHIKATA_KINGAKU],       
                        row[KARIKATA_KAMOKU], row[KARIKATA_HOJO_KAMOKU],
                        row[KASHIKATA_KAMOKU], row[KASHIKATA_HOJO_KAMOKU],
                        row[TEKIYOU]
                        )
                        # format(datetime.strptime(str(row[HIZUKE]), '%Y-%m-%d'),
                e.eprint("出納帳データ異常", str_msg)

    index_pair_set = set(index_pair_list)
    index_drop_list = []
    for pair in index_pair_set:
        # TODO 摘要    異なっていれば、合成する？
        index_drop_list.append(pair[1])
    df_ketsugou.drop(index_drop_list, inplace=True)
    df_ketsugou[DENPYOU_BANGOU] = df_ketsugou.index + 1

    # 複合仕訳と出納帳の重複を削除
    index_pair_list = []
    for _, row in df_fukugou_shiwake.iterrows():
        if \
                ((row[KARIKATA_KAMOKU], row[KARIKATA_HOJO_KAMOKU]) \
                        in list_suitou_kamoku) \
                and \
                (row[KASHIKATA_KAMOKU] == SHOKUCHI):
            # 借方が出納帳科目、貸方が諸口
            # 日付、借方、貸方、金額が同じものをリストアップ
            index_pair = df_ketsugou.index[ \
                    (df_ketsugou[HIZUKE] == row[HIZUKE]) \
                    & (df_ketsugou[KARIKATA_KAMOKU] == row[KARIKATA_KAMOKU]) \
                    & (df_ketsugou[KARIKATA_HOJO_KAMOKU] == row[KARIKATA_HOJO_KAMOKU]) \
                    & (df_ketsugou[KASHIKATA_KAMOKU] == SHOKUCHI) \
                    & (df_ketsugou[KARIKATA_KINGAKU] == row[KARIKATA_KINGAKU]) \
                    & (df_ketsugou[KASHIKATA_KINGAKU] == row[KASHIKATA_KINGAKU])]
            if (len(index_pair) > 0):
                # 一致したものの最初の１つだけを削除
                index_drop_list = index_pair[0:1]
                df_ketsugou.drop(index_drop_list, inplace=True)
        if \
                ((row[KASHIKATA_KAMOKU], row[KASHIKATA_HOJO_KAMOKU]) \
                        in list_suitou_kamoku) \
                and \
                (row[KARIKATA_KAMOKU] == SHOKUCHI):
            # 貸方が出納帳科目、借方が諸口
            # 日付、借方、貸方、金額が同じものをリストアップ
            index_pair = df_ketsugou.index[ \
                    (df_ketsugou[HIZUKE] == row[HIZUKE]) \
                    & (df_ketsugou[KASHIKATA_KAMOKU] == row[KASHIKATA_KAMOKU]) \
                    & (df_ketsugou[KASHIKATA_HOJO_KAMOKU] == row[KASHIKATA_HOJO_KAMOKU]) \
                    & (df_ketsugou[KARIKATA_KAMOKU] == SHOKUCHI) \
                    & (df_ketsugou[KARIKATA_KINGAKU] == row[KARIKATA_KINGAKU]) \
                    & (df_ketsugou[KASHIKATA_KINGAKU] == row[KASHIKATA_KINGAKU])]
            if (len(index_pair) > 0):
                # 一致したものの最初の１つだけを削除
                index_drop_list = index_pair[0:1]
                df_ketsugou.drop(index_drop_list, inplace=True)

    df_ketsugou = pd.concat([df_ketsugou, df_fukugou_shiwake])
    df_ketsugou.sort_values([HIZUKE],
            inplace=True)
    # d.dprint(df_ketsugou)
    # d.dprint_method_end()
    return df_ketsugou

def sakusei_soukanjou_motochou(shiwake_chou, kamoku, kishu_bi, kimatsu_bi):
    '''
    仕訳データから、指定された勘定科目の総勘定元帳データを作成する。

    Parameters
    ----------
    shiwake_chou : DataFrame
        仕訳データ。
    kamoku : tuple of str, int, boolean
        勘定科目名、期首残高、貸借区分のタプル
    kishu_bi : str
        期首日
    kimatsu_bi : str
        期末日

    Returns
    df_motochou : DataFrame
        総勘定元帳データ。
    zandaka : int
        期末残高。
    karikata_goukei : int
        借方金額の合計。
    kashikata_goukei : int
        貸方金額の合計。
    -------
    '''
    # d.dprint_method_start()
    # 借方
    df_karikata = shiwake_chou[ \
            shiwake_chou[KARIKATA_KAMOKU]==kamoku[0]].copy()
    df_karikata.loc[:, KASHIKATA_KINGAKU] = 0

    df_karikata_new = df_karikata.rename(columns= { \
            KARIKATA_HOJO_KAMOKU : HOJO_KAMOKU, \
            KASHIKATA_KAMOKU : AITE_KAMOKU, \
            KASHIKATA_HOJO_KAMOKU : AITE_HOJO_KAMOKU})
    del df_karikata
    df_karikata = df_karikata_new
    df_karikata = df_karikata.reindex([HIZUKE, DENPYOU_BANGOU,
            HOJO_KAMOKU, AITE_KAMOKU, AITE_HOJO_KAMOKU,
            TEKIYOU,
            KARIKATA_KINGAKU, KASHIKATA_KINGAKU, ZANDAKA],
            axis='columns')

    # 貸方
    df_kashikata = shiwake_chou[ \
            shiwake_chou[KASHIKATA_KAMOKU]==kamoku[0]].copy()
    df_kashikata.loc[:, KARIKATA_KINGAKU] = 0
    df_kashikata_new = df_kashikata.rename(columns= { \
            KASHIKATA_HOJO_KAMOKU : HOJO_KAMOKU, \
            KARIKATA_KAMOKU : AITE_KAMOKU, \
            KARIKATA_HOJO_KAMOKU : AITE_HOJO_KAMOKU})
    del df_kashikata
    df_kashikata = df_kashikata_new

    df_motochou = pd.concat([
            df_karikata.reindex([
                HIZUKE, DENPYOU_BANGOU,
                HOJO_KAMOKU, AITE_KAMOKU, AITE_HOJO_KAMOKU,
                TEKIYOU,
                KARIKATA_KINGAKU, KASHIKATA_KINGAKU, ZANDAKA],
                axis='columns'),
            df_kashikata.reindex([
                HIZUKE, DENPYOU_BANGOU,
                HOJO_KAMOKU, AITE_KAMOKU, AITE_HOJO_KAMOKU,
                TEKIYOU,
                KARIKATA_KINGAKU, KASHIKATA_KINGAKU, ZANDAKA],
                axis='columns')
            ])
    del df_karikata
    del df_kashikata
    df_motochou.sort_values([HIZUKE, DENPYOU_BANGOU], inplace=True)

    # 期首残高設定、残高計算、期末残高
    df_kishu = pd.DataFrame({
            HIZUKE : kishu_bi,
            DENPYOU_BANGOU : '',
            HOJO_KAMOKU : '',
            AITE_KAMOKU : '',
            AITE_HOJO_KAMOKU : '',
            TEKIYOU : '期首',
            KARIKATA_KINGAKU : '',
            KASHIKATA_KINGAKU : '',
            ZANDAKA : kamoku[1]
            }, index=[0])
    karikata_goukei = df_motochou[KARIKATA_KINGAKU].sum()
    kashikata_goukei = df_motochou[KASHIKATA_KINGAKU].sum()
    df_kimatsu = pd.DataFrame({
            HIZUKE : kimatsu_bi,
            DENPYOU_BANGOU : '',
            HOJO_KAMOKU : '',
            AITE_KAMOKU : '',
            AITE_HOJO_KAMOKU : '',
            TEKIYOU : '期末',
            KARIKATA_KINGAKU : karikata_goukei,
            KASHIKATA_KINGAKU : kashikata_goukei,
            ZANDAKA : 0
            }, index=[len(df_motochou)+1])
    df_new = pd.concat([df_kishu, df_motochou, df_kimatsu])
    del df_kishu
    del df_kimatsu
    del df_motochou
    df_motochou = df_new.reset_index(drop=True)
#     d.dprint(df_motochou)

    zandaka = kamoku[1]
    taishaku = kamoku[3]
    last = len(df_motochou) - 1
    for index, row in df_motochou.iterrows():
        if index == 0:
            continue
        if index != last:
            if taishaku:
                zandaka += row[KARIKATA_KINGAKU] - row[KASHIKATA_KINGAKU]
            else:
                zandaka += - row[KARIKATA_KINGAKU] + row[KASHIKATA_KINGAKU]
        df_motochou.at[index, ZANDAKA] = zandaka

    df_motochou[[ZANDAKA]] \
            = df_motochou[[ZANDAKA]].astype('int')

    # d.dprint_method_end()
    return df_motochou, zandaka, karikata_goukei, kashikata_goukei


def sakusei_hojo_motochou(shiwake_chou, hojo_kamoku,
        kishu_bi, kimatsu_bi):
    '''
    仕訳データから、指定された補助科目の補助元帳データを作成する。

    Parameters
    ----------
    shiwake_chou : DataFrame
        仕訳データ。
    hojo_kamoku : tuple of str, str, int, boolean
        勘定科目名、補助科目名、期首残高、貸借のタプル
    kishu_bi : str
        期首日
    kimatsu_bi : str
        期末日

    Returns
    df_motochou : DataFrame
        補助元帳データ。
    zandaka : int
        期末残高。
    karikata_goukei : int
        借方金額の合計。
    kashikata_goukei : int
        貸方金額の合計。
    -------
    '''
    d.dprint_method_start()
    d.dprint_name("hojo_kamoku", hojo_kamoku)
    # 借方
    df_karikata = shiwake_chou[ \
            (shiwake_chou[KARIKATA_KAMOKU]==hojo_kamoku[0]) \
            & (shiwake_chou[KARIKATA_HOJO_KAMOKU]==hojo_kamoku[1])] \
            .copy()
    df_karikata.loc[:, KASHIKATA_KINGAKU] = 0

    df_karikata_new = df_karikata.rename(columns= { \
#             KARIKATA_HOJO_KAMOKU : HOJO_KAMOKU, \
            KASHIKATA_KAMOKU : AITE_KAMOKU, \
            KASHIKATA_HOJO_KAMOKU : AITE_HOJO_KAMOKU})
    del df_karikata
    df_karikata = df_karikata_new
    df_karikata = df_karikata.reindex([HIZUKE, DENPYOU_BANGOU,
            AITE_KAMOKU, AITE_HOJO_KAMOKU,
            TEKIYOU,
            KARIKATA_KINGAKU, KASHIKATA_KINGAKU, ZANDAKA],
            axis='columns')

    # 貸方
    df_kashikata = shiwake_chou[ \
            (shiwake_chou[KASHIKATA_KAMOKU]==hojo_kamoku[0]) \
            & (shiwake_chou[KASHIKATA_HOJO_KAMOKU]==hojo_kamoku[1])] \
            .copy()
    df_kashikata.loc[:, KARIKATA_KINGAKU] = 0
    df_kashikata_new = df_kashikata.rename(columns= { \
#             KASHIKATA_HOJO_KAMOKU : HOJO_KAMOKU, \
            KARIKATA_KAMOKU : AITE_KAMOKU, \
            KARIKATA_HOJO_KAMOKU : AITE_HOJO_KAMOKU})
    del df_kashikata
    df_kashikata = df_kashikata_new

    df_motochou = pd.concat([
            df_karikata.reindex([
                HIZUKE, DENPYOU_BANGOU,
                AITE_KAMOKU, AITE_HOJO_KAMOKU,
                TEKIYOU,
                KARIKATA_KINGAKU, KASHIKATA_KINGAKU, ZANDAKA],
                axis='columns'),
            df_kashikata.reindex([
                HIZUKE, DENPYOU_BANGOU,
                AITE_KAMOKU, AITE_HOJO_KAMOKU,
                TEKIYOU,
                KARIKATA_KINGAKU, KASHIKATA_KINGAKU, ZANDAKA],
                axis='columns')
            ])
    del df_karikata
    del df_kashikata
    df_motochou.sort_values([HIZUKE, DENPYOU_BANGOU], inplace=True)

    # 期首残高設定、残高計算、期末残高
    df_kishu = pd.DataFrame({
            HIZUKE : kishu_bi,
            DENPYOU_BANGOU : '',
            AITE_KAMOKU : '',
            AITE_HOJO_KAMOKU : '',
            TEKIYOU : '期首',
            KARIKATA_KINGAKU : '',
            KASHIKATA_KINGAKU : '',
            ZANDAKA : hojo_kamoku[2]
            }, index=[0])
    karikata_goukei = df_motochou[KARIKATA_KINGAKU].sum()
    kashikata_goukei = df_motochou[KASHIKATA_KINGAKU].sum()
    df_kimatsu = pd.DataFrame({
            HIZUKE : kimatsu_bi,
            DENPYOU_BANGOU : '',
            AITE_KAMOKU : '',
            AITE_HOJO_KAMOKU : '',
            TEKIYOU : '期末',
            KARIKATA_KINGAKU : karikata_goukei,
            KASHIKATA_KINGAKU : kashikata_goukei,
            ZANDAKA : 0
            }, index=[len(df_motochou)+1])
    df_new = pd.concat([df_kishu, df_motochou, df_kimatsu])
    del df_kishu
    del df_kimatsu
    del df_motochou
    df_motochou = df_new.reset_index(drop=True)

    zandaka = hojo_kamoku[2]
    taishaku = hojo_kamoku[4]
    last = len(df_motochou) - 1
    for index, row in df_motochou.iterrows():
        if index == 0:
            continue
        if index != last:
            if taishaku:
                zandaka += row[KARIKATA_KINGAKU] - row[KASHIKATA_KINGAKU]
            else:
                zandaka += - row[KARIKATA_KINGAKU] + row[KASHIKATA_KINGAKU]
        df_motochou.at[index, ZANDAKA] = zandaka

    df_motochou[[ZANDAKA]] \
            = df_motochou[[ZANDAKA]].astype('int')

    d.dprint_method_end()
    return df_motochou, zandaka, karikata_goukei, kashikata_goukei


def henkan_taishaku_kubun(taishaku_kubun):
    '''
    貸借区分（資産・負債・収入・支出）から、貸借（True,Fasle）へ変換する。

    Parameters
    ----------
    taishaku_kubun : str
        貸借区分（資産・負債・収入・支出）。

    Returns
    taishaku : boolean
        貸借（借方-True、貸方-False）。
    -------
    '''
    d.dprint_method_start()
    if (taishaku_kubun == TAISHAKU_KUBUN_KARI) \
            or (taishaku_kubun == TAISHAKU_KUBUN_SHISAN) \
            or (taishaku_kubun == TAISHAKU_KUBUN_SHISHUTSU) \
            or (taishaku_kubun == TAISHAKU_KUBUN_HIYOU):
        d.dprint_method_end()
        return True
    if (taishaku_kubun == TAISHAKU_KUBUN_KASHI) \
            or (taishaku_kubun == TAISHAKU_KUBUN_JUNSHISAN) \
            or (taishaku_kubun == TAISHAKU_KUBUN_FUSAI) \
            or (taishaku_kubun == TAISHAKU_KUBUN_SHUNYU) \
            or (taishaku_kubun == TAISHAKU_KUBUN_SHIHON) \
            or (taishaku_kubun == TAISHAKU_KUBUN_SHUEKI):
        d.dprint_method_end()
        return False
    e.eprint("貸借区分が誤っている。",
            "誤った貸借区分「{}」".format(taishaku_kubun))
    d.dprint_method_end()
    return True


SIDE_NORMAL = Side(border_style="thin", color='000000')
BORDER_NORMAL = Border(top=SIDE_NORMAL, left=SIDE_NORMAL,
        right=SIDE_NORMAL, bottom=SIDE_NORMAL)

def save_soukanjou_motochou_file(file_name,
        dantai_mei, kaishi_bi, kimatsu_bi,
        soukanjou_motochou_list):
    '''
    総勘定元帳データをExcelファイルに保存する。

    Parameters
    ----------
    file_name : str
        保存するファイル名。
    dantai_mei : str
        団体名。
    kishu_bi : str
        期首日。
    kimatsu_bi : str
        期末日
    soukanjou_motochou_list : list of tuple
            of (tuple of str, int, str), DataFrame
        勘定科目、期首残高、貸借区分のタプル、総勘定元帳データのタプルのリスト

    Returns
    -------
    '''
    d.dprint_method_start()
    wb = xl.Workbook()
    del wb['Sheet']
    for (kamoku, motochou) in soukanjou_motochou_list:
        sheet = wb.create_sheet(title=kamoku[0])
        sheet["D1"] = "総勘定元帳"
        sheet["F1"] = kamoku[0] # 勘定科目名
        sheet["H1"] = kamoku[2] # 資産、収入など
        sheet["A2"] = dantai_mei
#         tstr = '1900-01-01 00:00:00'
#         tdatetime = datetime.strptime(tstr, '%Y-%m-%d %H:%M:%S')
#         d.dprint(tdatetime.toordinal())

        str_kishu = "TEXT(" \
                + str(kaishi_bi.toordinal()-693594) \
                + ',"' + FORMAT_KIKAN + '")'
        str_kimatsu = "TEXT(" \
                + str(kimatsu_bi.toordinal()-693594) \
                + ',"' + FORMAT_KIKAN + '")'
        sheet["F2"] = "=CONCATENATE(" + str_kishu \
                + ',"～" ,' + str_kimatsu + ")"
        sheet["I2"] = TANI_EN
        sheet["I2"].alignment \
                = Alignment(horizontal='right',vertical='center')

        rows = dataframe_to_rows(
                motochou, index=False, header=True)
        for row in rows:
            sheet.append(row)

        sheet.print_title_rows = '1:3'
        sheet.oddFooter.center.text = "Page &[Page] of &N"
        sheet.page_setup.orientation \
                = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

        global HIZUKE_W, BANGOU_W
        global KAMOKU_W, HOJO_W, KINGAKU_W
        global TEKIYOU1_W, TEKIYOU2_W, TEKIYOU_W
        sheet.column_dimensions['A'].width = HIZUKE_W
        sheet.column_dimensions['B'].width = BANGOU_W
        sheet.column_dimensions['C'].width = HOJO_W
        sheet.column_dimensions['D'].width = KAMOKU_W
        sheet.column_dimensions['E'].width = HOJO_W
        sheet.column_dimensions['F'].width = TEKIYOU_W
        sheet.column_dimensions['G'].width = KINGAKU_W
        sheet.column_dimensions['H'].width = KINGAKU_W
        sheet.column_dimensions['I'].width = KINGAKU_W

        for row_index, row in enumerate(sheet):
            sheet.row_dimensions[row_index + 1].height = TAKASA
            if row_index < 2:
                continue
            if row_index == 2:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    sheet[cell.coordinate].border = BORDER_NORMAL
                continue
            for cell_index, cell in enumerate(row):
                if cell_index == 0:
#                     cell.number_format = u'yyyy年m月d日'
                    cell.number_format = FORMAT_HIZUKE
                elif cell_index == 6 \
                        or cell_index == 7 \
                        or cell_index == 8:
                    cell.number_format = "#,##0"
                sheet[cell.coordinate].border = BORDER_NORMAL

    try:
        wb.save(file_name)
    except Exception as error:
        e.eprint('Ｅｘｃｅｌ保存エラー', error)
    d.dprint_method_end()
    return


def save_hojo_motochou_file(file_name,
        dantai_mei, kaishi_bi, kimatsu_bi,
        hojo_motochou_list):
    '''
    補助元帳データをExcelファイルに保存する。

    Parameters
    ----------
    file_name : str
        保存するファイル名。
    dantai_mei : str
        団体名。
    kishu_bi : str
        期首日。
    kimatsu_bi : str
        期末日
    hojo_motochou_list : list of tuple
            of (tuple of str, int, str), DataFrame
        勘定科目、補助科目、期首残高、貸借区分のタプル、補助定元帳データのタプルのリスト

    Returns
    -------
    '''
    d.dprint_method_start()
    wb = xl.Workbook()
    del wb['Sheet']
    for (hojo_kamoku, motochou) in hojo_motochou_list:
        sheet = wb.create_sheet(title=hojo_kamoku[0] + '_' + hojo_kamoku[1])
        sheet["C1"] = "補助元帳"
        sheet["E1"] = hojo_kamoku[0] + '　' + hojo_kamoku[1] # 勘定科目名　補助科目名
#         sheet["G1"] = kamoku[3] # 資産、収入など
        sheet["G1"] = hojo_kamoku[3] # 資産、収入など
        sheet["A2"] = dantai_mei
        str_kishu = "TEXT(" \
                + str(kaishi_bi.toordinal()-693594) \
                + ',"' + FORMAT_KIKAN + '")'
        str_kimatsu = "TEXT(" \
                + str(kimatsu_bi.toordinal()-693594) \
                + ',"' + FORMAT_KIKAN + '")'
        sheet["E2"] = "=CONCATENATE(" + str_kishu \
                + ',"～" ,' + str_kimatsu + ")"
#         sheet["E2"] = str(kaishi_bi.year) +"年" \
#                 + str(kaishi_bi.month) + "月" \
#                 + str(kaishi_bi.day) + "日　～" \
#                 + str(kimatsu_bi.year) +"年" \
#                 + str(kimatsu_bi.month) + "月" \
#                 + str(kimatsu_bi.day) + "日"
        sheet["H2"] = TANI_EN
        sheet["H2"].alignment \
                = Alignment(horizontal='right',vertical='center')

        rows = dataframe_to_rows(
                motochou, index=False, header=True)
        for row in rows:
            sheet.append(row)

        sheet.print_title_rows = '1:3'
        sheet.oddFooter.center.text = "Page &[Page] of &N"
        sheet.page_setup.orientation \
                = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

        global HIZUKE_W, BANGOU_W
        global KAMOKU_W, HOJO_W, KINGAKU_W
        global TEKIYOU1_W, TEKIYOU2_W, TEKIYOU_W
        sheet.column_dimensions['A'].width = HIZUKE_W
        sheet.column_dimensions['B'].width = BANGOU_W
        sheet.column_dimensions['C'].width = KAMOKU_W
        sheet.column_dimensions['D'].width = HOJO_W
        sheet.column_dimensions['E'].width = TEKIYOU_W
        sheet.column_dimensions['F'].width = KINGAKU_W
        sheet.column_dimensions['G'].width = KINGAKU_W
        sheet.column_dimensions['H'].width = KINGAKU_W

        for row_index, row in enumerate(sheet):
            sheet.row_dimensions[row_index + 1].height = TAKASA
            if row_index < 2:
                continue
            if row_index == 2:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    sheet[cell.coordinate].border = BORDER_NORMAL
                continue
            for cell_index, cell in enumerate(row):
                if cell_index == 0:
#                     cell.number_format = u'yyyy年m月d日'
                    cell.number_format = FORMAT_HIZUKE
                elif cell_index == 5 \
                        or cell_index == 6 \
                        or cell_index == 7:
                    cell.number_format = "#,##0"
                sheet[cell.coordinate].border = BORDER_NORMAL

    try:
        wb.save(file_name)
    except Exception as error:
        e.eprint('Ｅｘｃｅｌ保存エラー', error)
    d.dprint_method_end()
    return


def save_shisanhyou_file(file_name,
        dantai_mei, kaishi_bi, kimatsu_bi,
        shisanhyou_list, hojo_ichiran_list):
    '''
    残高試算表、補助残高一覧表データをExcelファイルに保存する。

    Parameters
    ----------
    file_name : str
        保存するファイル名。
    dantai_mei : str
        団体名。
    kishu_bi : str
        期首日。
    kimatsu_bi : str
        期末日。
    shisanhyou_list : list of tuple of str, int, int, int, int
        勘定科目、期首残高、借方金額、貸方金額、期末残高のタプルのリスト
    hojo_ichiran_list : list of tuple of str, str, int, int, int, int
        勘定科目、補助科目、期首残高、借方金額、貸方金額、期末残高のタプルのリスト

    Returns
    -------
    '''
    d.dprint_method_start()
    wb = xl.Workbook()
    del wb['Sheet']
    sheet = wb.create_sheet(title="残高試算表")
    sheet["A1"] = "残高試算表"
    sheet["A1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["B1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["C1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["D1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["E1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["E2"] = TANI_EN
    sheet["E2"].alignment \
            = Alignment(horizontal='right',vertical='center')


    sheet["A2"] = dantai_mei
    str_kishu = "TEXT(" \
            + str(kaishi_bi.toordinal()-693594) \
            + ',"' + FORMAT_KIKAN + '")'
    str_kimatsu = "TEXT(" \
            + str(kimatsu_bi.toordinal()-693594) \
            + ',"' + FORMAT_KIKAN + '")'
    sheet["B2"] = "=CONCATENATE(" + str_kishu \
            + ',"～" ,' + str_kimatsu + ")"
#     sheet["B2"] = str(kaishi_bi.year) +"年" \
#             + str(kaishi_bi.month) + "月" \
#             + str(kaishi_bi.day) + "日　～　" \
#             + str(kimatsu_bi.year) +"年" \
#             + str(kimatsu_bi.month) + "月" \
#             + str(kimatsu_bi.day) + "日"
    sheet["B2"].alignment = Alignment(horizontal="left")
    sheet["A3"] = "勘定科目名"
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet["B3"] = "期首残高"
    sheet["B3"].alignment = Alignment(horizontal="center")
    sheet["C3"] = "借方金額"
    sheet["C3"].alignment = Alignment(horizontal="center")
    sheet["D3"] = "貸方金額"
    sheet["D3"].alignment = Alignment(horizontal="center")
    sheet["E3"] = "期末残高"
    sheet["E3"].alignment = Alignment(horizontal="center")

    for kamoku in shisanhyou_list:
        sheet.append(kamoku)

    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    global HIZUKE_W, BANGOU_W
    global KAMOKU_W, HOJO_W, KINGAKU_W
    global TEKIYOU1_W, TEKIYOU2_W, TEKIYOU_W

    sheet.column_dimensions['A'].width = KAMOKU_W
    sheet.column_dimensions['B'].width = KINGAKU_W
    sheet.column_dimensions['C'].width = KINGAKU_W
    sheet.column_dimensions['D'].width = KINGAKU_W
    sheet.column_dimensions['E'].width = KINGAKU_W

    for row_index, row in enumerate(sheet):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 2:
            continue
        for cell in row:
            cell.number_format = "#,##0"
            sheet[cell.coordinate].border = BORDER_NORMAL

    sheet = wb.create_sheet(title="補助残高一覧表")
    sheet["A1"] = "補助残高一覧表"
    sheet["A1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["B1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["C1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["D1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["E1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["F1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["F2"] = TANI_EN
    sheet["F2"].alignment \
            = Alignment(horizontal='right',vertical='center')

    sheet["A2"] = dantai_mei
    str_kishu = "TEXT(" \
            + str(kaishi_bi.toordinal()-693594) \
            + ',"' + FORMAT_KIKAN + '")'
    str_kimatsu = "TEXT(" \
            + str(kimatsu_bi.toordinal()-693594) \
            + ',"' + FORMAT_KIKAN + '")'
    sheet["B2"] = "=CONCATENATE(" + str_kishu \
            + ',"～" ,' + str_kimatsu + ")"
#     sheet["B2"] = str(kaishi_bi.year) +"年" \
#             + str(kaishi_bi.month) + "月" \
#             + str(kaishi_bi.day) + "日　～　" \
#             + str(kimatsu_bi.year) +"年" \
#             + str(kimatsu_bi.month) + "月" \
#             + str(kimatsu_bi.day) + "日"
    sheet["B2"].alignment = Alignment(horizontal="left")
    sheet["A3"] = "勘定科目名"
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet["B3"] = "補助科目名"
    sheet["B3"].alignment = Alignment(horizontal="center")
    sheet["C3"] = "期首残高"
    sheet["C3"].alignment = Alignment(horizontal="center")
    sheet["D3"] = "借方金額"
    sheet["D3"].alignment = Alignment(horizontal="center")
    sheet["E3"] = "貸方金額"
    sheet["E3"].alignment = Alignment(horizontal="center")
    sheet["F3"] = "期末残高"
    sheet["F3"].alignment = Alignment(horizontal="center")

    for hojo_kamoku in hojo_ichiran_list:
        sheet.append(hojo_kamoku)

    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    sheet.column_dimensions['A'].width = KAMOKU_W
    sheet.column_dimensions['B'].width = HOJO_W
    sheet.column_dimensions['C'].width = KINGAKU_W
    sheet.column_dimensions['D'].width = KINGAKU_W
    sheet.column_dimensions['E'].width = KINGAKU_W
    sheet.column_dimensions['G'].width = KINGAKU_W

    for row_index, row in enumerate(sheet):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 2:
            continue
        for cell in row:
            cell.number_format = "#,##0"
            sheet[cell.coordinate].border = BORDER_NORMAL

    try:
        wb.save(file_name)
    except Exception as error:
        e.eprint('Ｅｘｃｅｌ保存エラー', error)
    d.dprint_method_end()
    return


def save_shiwakechou_file(file_name,
        dantai_mei, kaishi_bi, kimatsu_bi,
        shiwakechou):
    '''
    仕訳データをExcelファイルに保存する。

    Parameters
    ----------
    file_name : str
        保存するファイル名。
    dantai_mei : str
        団体名。
    kishu_bi : str
        期首日。
    kimatsu_bi : str
        期末日。
    shiwakechou : DataFrame
        仕訳データ。

    Returns
    -------
    '''
    d.dprint_method_start()
    wb = xl.Workbook()
    del wb['Sheet']
    sheet = wb.create_sheet(title="仕訳帳")
    sheet["A1"] = "仕訳帳"
    sheet["A1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["B1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["C1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["D1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["E1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["F1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["G1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["H1"].alignment = Alignment(horizontal="centerContinuous")
    sheet["I1"].alignment = Alignment(horizontal="centerContinuous")

    sheet["A2"] = dantai_mei
    str_kishu = "TEXT(" \
            + str(kaishi_bi.toordinal()-693594) \
            + ',"' + FORMAT_KIKAN + '")'
    str_kimatsu = "TEXT(" \
            + str(kimatsu_bi.toordinal()-693594) \
            + ',"' + FORMAT_KIKAN + '")'
    # sheet["B2"] = "=CONCATENATE(" + str_kishu \
    #         + ',"～" ,' + str_kimatsu + ")"
    # version0.11
    sheet["E2"] = "=CONCATENATE(" + str_kishu \
            + ',"～" ,' + str_kimatsu + ")"
#     sheet["B2"] = str(kaishi_bi.year) +"年" \
#             + str(kaishi_bi.month) + "月" \
#             + str(kaishi_bi.day) + "日　～　" \
#             + str(kimatsu_bi.year) +"年" \
#             + str(kimatsu_bi.month) + "月" \
#             + str(kimatsu_bi.day) + "日"
    sheet["B2"].alignment = Alignment(horizontal="left")
    sheet["A3"] = "日付"
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet["B3"] = "番号"
    sheet["B3"].alignment = Alignment(horizontal="center")
    sheet["C3"] = "借方科目"
    sheet["C3"].alignment = Alignment(horizontal="center")
    sheet["D3"] = "借方補助"    # "借方補助科目"
    sheet["D3"].alignment = Alignment(horizontal="center")
    sheet["E3"] = "借方金額"
    sheet["E3"].alignment = Alignment(horizontal="center")
    sheet["F3"] = "貸方科目"
    sheet["F3"].alignment = Alignment(horizontal="center")
    sheet["G3"] = "貸方補助"  # "貸方補助科目"
    sheet["G3"].alignment = Alignment(horizontal="center")
    sheet["H3"] = "貸方金額"
    sheet["H3"].alignment = Alignment(horizontal="center")
    sheet["I3"] = "摘要"
    sheet["I3"].alignment = Alignment(horizontal="center")
    sheet["I2"] = TANI_EN
    sheet["I2"].alignment \
            = Alignment(horizontal='right',vertical='center')

    rows = dataframe_to_rows(
            shiwakechou, index=False, header=False)
    for row in rows:
        sheet.append(row)

    sheet.print_title_rows = '1:3'
    sheet.oddFooter.center.text = "Page &[Page] of &N"
    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    global HIZUKE_W, BANGOU_W
    global KAMOKU_W, HOJO_W, KINGAKU_W
    global TEKIYOU1_W, TEKIYOU2_W, TEKIYOU_W
    sheet.column_dimensions['A'].width = HIZUKE_W
    sheet.column_dimensions['B'].width = BANGOU_W
    sheet.column_dimensions['C'].width = KAMOKU_W
    sheet.column_dimensions['D'].width = HOJO_W
    sheet.column_dimensions['E'].width = KINGAKU_W
    sheet.column_dimensions['F'].width = KAMOKU_W
    sheet.column_dimensions['G'].width = HOJO_W
    sheet.column_dimensions['H'].width = KINGAKU_W
    sheet.column_dimensions['I'].width = TEKIYOU_W

    for row_index, row in enumerate(sheet):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 2:
            continue
        if row_index == 2:
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                sheet[cell.coordinate].border = BORDER_NORMAL
            continue
        for cell_index, cell in enumerate(row):
            if cell_index == 0:
#                 cell.number_format = u'yyyy年m月d日'
                cell.number_format = FORMAT_HIZUKE
            elif cell_index == 4 \
                    or cell_index == 7:
                cell.number_format = "#,##0"
            sheet[cell.coordinate].border = BORDER_NORMAL

    try:
        wb.save(file_name)
    except Exception as error:
        e.eprint('Ｅｘｃｅｌ保存エラー', error)
    d.dprint_method_end()
    return


def save_yokuki_kihon(file_name,
            dantai_mei, kishu_bi, kimatsu_bi,
            suitou_list, shisanhyou_list, hojo_ichiran_list,
            kamoku_list, hojo_kamoku_list):
    '''
    翌期用のExcelファイルを作成する。

    Parameters
    ----------
    file_name : str
        保存するファイル名。
    dantai_mei : str
        団体名。
    kishu_bi : str
        当期の期首日。
    kimatsu_bi : str
        当期の期末日。
    suitou_list : list of tuple of ...
        出納データがある勘定科目、補助科目
    shisanhyou_list : list of tuple of str, int, int, int, int
        勘定科目、期首残高、借方金額、貸方金額、期末残高のタプルのリスト
    hojo_ichiran_list : list of tuple of str, str, int, int, int, int
        勘定科目、補助科目、期首残高、借方金額、貸方金額、期末残高のタプルのリスト
    kamoku_list : list of tuple of str,...

    hojo_kamoku_list : list of tuple of ...

    Returns
    -------
    '''
    d.dprint_method_start()
    wb = xl.Workbook()
    del wb['Sheet']
    (row_start, row_end) = create_yokuki_kihon_sheet(wb,
            dantai_mei, kishu_bi, kimatsu_bi,
            suitou_list, shisanhyou_list, hojo_ichiran_list,
            kamoku_list, hojo_kamoku_list)

    kamoku_unique_list = []
    for kamoku in kamoku_list:
        kamoku_unique_list.append(kamoku[0])
    str_kamoku_list = '"' + ','.join(kamoku_unique_list) + '"'
    dv_kamoku = DataValidation(type="list",
            formula1=str_kamoku_list, allow_blank=True)
    dv_kamoku_tanitsu = DataValidation(type="list",
            formula1=str_kamoku_list, allow_blank=True)

    hojo_unique_list = []
    for hojo in hojo_kamoku_list:
        if hojo[1] not in hojo_unique_list:
            hojo_unique_list.append(hojo[1])
    str_hojo_list = '"' + ','.join(hojo_unique_list) + '"'
    dv_hojo = DataValidation(type="list",
            formula1=str_hojo_list, allow_blank=True)
    dv_hojo_tanitsu = DataValidation(type="list",
            formula1=str_hojo_list, allow_blank=True)

    # 出納帳用のシート作成
    for suitou in suitou_list:
        if len(suitou) == 1:
            for kamoku in shisanhyou_list:
                if kamoku[0] == suitou[0]:
                    break
            create_yokuki_suitou_sheet(wb,
                    suitou[0],
                    kishu_bi,
                    suitou[0], kamoku[4],
                    dv_kamoku, dv_hojo)
        elif len(suitou) == 2:
            for hojo in hojo_ichiran_list:
                if (hojo[0] == suitou[0]) \
                        and (hojo[1] == suitou[1]):
                    break
            # version0.11
            create_yokuki_suitou_sheet(wb,
                    suitou[0] + '_' + suitou[1],
                    kishu_bi,
                    suitou[0] + ' ' + suitou[1], hojo[5],
                    dv_kamoku, dv_hojo)

    # 単一仕訳用のシート作成
    create_yokuki_tanitsushiwake_sheet(wb, TANITSU_SHEET_NAME,
            dv_kamoku_tanitsu, dv_hojo_tanitsu)
    # 複合仕訳用のシート作成
    create_yokuki_fukugoushiwake_sheet(wb, FUKUGOU_SHEET_NAME,
            dv_kamoku_tanitsu, dv_hojo_tanitsu)

    try:
        wb.save(file_name)
    except Exception as error:
        e.eprint('Ｅｘｃｅｌ保存エラー', error)
    d.dprint_method_end()
    return

def create_yokuki_kihon_sheet(wb,
            dantai_mei, kishu_bi, kimatsu_bi,
            suitou_list, shisanhyou_list, hojo_ichiran_list,
            kamoku_list, hojo_kamoku_list):
    sheet = wb.create_sheet(title="基本")
    sheet["A1"] = "団体名"
    sheet['B1'] = dantai_mei
    sheet['A2'] = '期首日'
#     yokuki_kishubi = kishu_bi + relativedelta(years=1)
#     sheet['B2'] = "=TEXT(" \
#             + str(yokuki_kishubi.toordinal()-693594) \
#             + ',"' + FORMAT_KIKAN + '")'
#     sheet["B2"].number_format = FORMAT_KIKAN
    sheet['B2'] = str(kishu_bi.year+1) +"-" \
            + str(kishu_bi.month) + "-" \
            + str(kishu_bi.day)
#     sheet["B2"].number_format = FORMAT_HIZUKE
    sheet['A3'] = '期末日'
#     yokuki_kimatsubi = kimatsu_bi + relativedelta(years=1)
#     sheet['B3'] = "=TEXT(" \
#                 + str(yokuki_kimatsubi.toordinal()-693594) \
#                 + ',"' + FORMAT_KIKAN + '")'
#     sheet["B3"].number_format = FORMAT_KIKAN
    sheet['B3'] = str(kimatsu_bi.year+1) +"-" \
            + str(kimatsu_bi.month) + "-" \
            + str(kimatsu_bi.day)
#     sheet["B3"].number_format = FORMAT_HIZUKE
    sheet["A4"] = "科目"
    sheet["B4"] = "補助科目"
    sheet["C4"] = "期首残高"
    sheet["D4"] = "補助別期首残高"
    sheet["E4"] = "出納"
    sheet["F4"] = "区分"

    hojo_set = set()
    for hojo in hojo_ichiran_list:
        hojo_set.add(hojo[0])   # 補助科目がある勘定科目を登録
    row_num = 0
    for kamoku in shisanhyou_list:
#         if (kamoku[0][0] == '+') or (kamoku[0][0] == '*'):
#             continue
        if kamoku[0] in hojo_set:
            # 補助科目の処理
            # for hojo in hojo_ichiran_list:
            for hojo in hojo_ichiran_list:
                if kamoku[0] == hojo[0]:
                    if hojo[1] != "【合計】":
                        suitou_umu = ''
                        for suitou in suitou_list:
                            if len(suitou) == 2:
                                if (suitou[0] == kamoku[0]) \
                                        and (suitou[1] == hojo[1]):
                                    suitou_umu = '有'
                        for hojo2 in hojo_kamoku_list:
                            if (hojo[0] == hojo2[0]) \
                                    and (hojo[1] == hojo2[1]):
                                kubun = hojo2[3]
                        if (kubun == TAISHAKU_KUBUN_SHUNYU) \
                                or (kubun == TAISHAKU_KUBUN_SHISHUTSU):
                            zandaka = 0
                        else:
                            zandaka = hojo[5]
                        sheet.append((kamoku[0], hojo[1], '', zandaka,
                                suitou_umu, kubun))
                        row_num += 1
            for kamoku2 in kamoku_list:
                if kamoku[0] == kamoku2[0]:
                    kubun = kamoku2[2]
            if (kubun == TAISHAKU_KUBUN_SHUNYU) \
                    or (kubun == TAISHAKU_KUBUN_SHISHUTSU):
                zandaka = 0
            else:
                zandaka = kamoku[4]
            sheet.append((kamoku[0], "【合計】", zandaka, '',
                    '', kubun))
            row_num += 1
        else:
            # 補助科目がない勘定科目の処理
            suitou_umu = ''
            for suitou in suitou_list:
                if len(suitou) == 1:
                    if suitou[0] == kamoku[0]:
                        suitou_umu = '有'
            for kamoku2 in kamoku_list:
                if kamoku[0] == kamoku2[0]:
                    kubun = kamoku2[2]
            if (kubun == TAISHAKU_KUBUN_SHUNYU) \
                    or (kubun == TAISHAKU_KUBUN_SHISHUTSU):
                zandaka = 0
            else:
                zandaka = kamoku[4]
            sheet.append((kamoku[0], '', zandaka, '',
                    suitou_umu, kubun))
            row_num += 1

    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    global KAMOKU_W, HOJO_W, KINGAKU_W
    global SUITOU_W, KUBUN_W
    global HIZUKE_W, TEKIYOU_W
    sheet.column_dimensions['A'].width = KAMOKU_W
    sheet.column_dimensions['B'].width = HOJO_W
    sheet.column_dimensions['C'].width = KINGAKU_W
    sheet.column_dimensions['D'].width = KINGAKU_W
    sheet.column_dimensions['E'].width = SUITOU_W
    sheet.column_dimensions['F'].width = KUBUN_W

    global TAKASA
    for row_index, row in enumerate(sheet):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 3:
            continue
        if row_index == 3:
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                sheet[cell.coordinate].border = BORDER_NORMAL
            continue
        for cell_index, cell in enumerate(row):
            if cell_index == 2 \
                    or cell_index == 3:
                cell.number_format = "#,##0"
            sheet[cell.coordinate].border = BORDER_NORMAL

    row_end = 5 + row_num
#     d.dprint(row_end)
    return (5, row_end)


def create_yokuki_suitou_sheet(wb, sheet_name, kishu_bi,
            kamoku_name, kishu_zandaka,
            dv_kamoku, dv_hojo):
    sheet = wb.create_sheet(title=sheet_name)
    sheet["A1"] = kamoku_name

    sheet["A2"] = HIZUKE
    sheet["B2"] = AITE_KAMOKU
    sheet["C2"] = AITE_HOJO_KAMOKU
    sheet["D2"] = TEKIYOU
    sheet["E2"] = TEKIYOU2
    sheet["F2"] = NYUKIN
    sheet["G2"] = SHUKKIN
    sheet["H2"] = ZANDAKA

    # 期首残高欄
    kishu_bi = kishu_bi + relativedelta(years=1)
    str_kishu = "=TEXT(" \
            + str(kishu_bi.toordinal()-693594) \
            + ',"' + FORMAT_HIZUKE + '")'
    sheet["A3"] = str_kishu
    sheet["A3"].number_format = FORMAT_HIZUKE
    # version0.11
    sheet["A3"].alignment \
            = Alignment(horizontal="right")
    sheet["H3"] = kishu_zandaka
    sheet["H3"].number_format = "#,##0"

    sheet["A3"].border = BORDER_NORMAL
    side_obj = Side(border_style='thin')
    border_obj = Border(diagonal=side_obj,
            diagonalDown=False, diagonalUp=True,
            top=SIDE_NORMAL, left=SIDE_NORMAL,
            right=SIDE_NORMAL, bottom=SIDE_NORMAL)
    sheet["B3"].border = border_obj
    sheet["C3"].border = border_obj
    sheet["D3"].border = border_obj
    sheet["E3"].border = border_obj
    sheet["F3"].border = border_obj
    sheet["G3"].border = border_obj
    sheet["H3"].border = BORDER_NORMAL

    global KAMOKU_W, HOJO_W, KINGAKU_W
    global SUITOU_W, KUBUN_W
    global HIZUKE_W, TEKIYOU_W
    global TAKASA
    for row_index in range(50):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 1:
            continue
        if row_index == 2:
            for column_index in range(1, 8+1):
                sheet.cell(row=row_index, column=column_index) \
                        .alignment = Alignment(horizontal="center")
        if row_index > 3:
            # str_shiki = "=H{}+F{}-G{}".format(row_index-1, row_index, row_index)
            str_shiki = "=OFFSET(H{},-1,0)+F{}-G{}".format(row_index, row_index, row_index)
            sheet.append(('', '', '', '', '', '', '', str_shiki))
            sheet.cell(row=row_index, column=1).number_format = FORMAT_HIZUKE
            sheet.cell(row=row_index, column=4).number_format = "@"
            sheet.cell(row=row_index, column=4).number_format = "@"
            sheet.cell(row=row_index, column=6).number_format = "#,##0"
            sheet.cell(row=row_index, column=7).number_format = "#,##0"
            sheet.cell(row=row_index, column=8).number_format = "#,##0"

            dv_kamoku.add(sheet.cell(row_index, 2))
            dv_hojo.add(sheet.cell(row_index, 3))

        if (row_index == 2) or (row_index > 3):
            for column_index in range(1, 8+1):
                sheet.cell(row=row_index, column=column_index) \
                        .border = BORDER_NORMAL
    sheet.column_dimensions['A'].width = HIZUKE_W
    sheet.column_dimensions['B'].width = KAMOKU_W
    sheet.column_dimensions['C'].width = HOJO_W
    sheet.column_dimensions['D'].width = TEKIYOU1_W
    sheet.column_dimensions['E'].width = TEKIYOU2_W
    sheet.column_dimensions['F'].width = KINGAKU_W
    sheet.column_dimensions['G'].width = KINGAKU_W
    sheet.column_dimensions['H'].width = KINGAKU_W

    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    sheet.add_data_validation(dv_kamoku)
    sheet.add_data_validation(dv_hojo)

    return

def create_yokuki_tanitsushiwake_sheet(wb, sheet_name,
            dv_kamoku, dv_hojo):
    '''
    翌期入力用のExcelファイルに単一仕訳シートを作成する。

    Parameters
    ----------
    wb : workbook
        翌期入力用Excelファイル
    sheet_name : sheet_name
        単一仕訳シートのシート名
    dv_kamoku :
        勘定科目のドロップダウンリスト設定用データ
    dv_hojo :
        補助科目のドロップダウンリスト設定用データ

    Returns
    -------
    '''
    sheet = wb.create_sheet(title=sheet_name)
    sheet["A1"] = sheet_name

    sheet["A2"] = HIZUKE
    sheet["B2"] = KARIKATA_KAMOKU
    sheet["C2"] = KARIKATA_HOJO_KAMOKU
    sheet["D2"] = KARIKATA_KINGAKU
    sheet["E2"] = KASHIKATA_KAMOKU
    sheet["F2"] = KASHIKATA_HOJO_KAMOKU
    sheet["G2"] = KASHIKATA_KINGAKU
    sheet["H2"] = TEKIYOU
    sheet["I2"] = TEKIYOU2

    global TAKASA
    for row_index in range(50):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 1:
            continue
        if row_index == 2:
            for column_index in range(1, 9+1):
                sheet.cell(row=row_index, column=column_index) \
                        .alignment = \
                        Alignment(horizontal="center")
        if row_index > 2:
            sheet.append(('', '', '', '', '', '', '', '', ''))
            # 書式の設定
            sheet.cell(row=row_index, column=1).number_format = FORMAT_HIZUKE
            sheet.cell(row=row_index, column=4).number_format = "#,##0"
            sheet.cell(row=row_index, column=7).number_format = "#,##0"
            # ドロップダウンリストの設定
            dv_kamoku.add(sheet.cell(row_index, 2))
            dv_hojo.add(sheet.cell(row_index, 3))
            dv_kamoku.add(sheet.cell(row_index, 5))
            dv_hojo.add(sheet.cell(row_index, 6))
            # 貸方金額に式を設定
            # sheet.cell(row=row_index, column=7).value = \
            #         '=RC[-3]'
            sheet.cell(row=row_index, column=7).value = \
                    '=D{}'.format(row_index)
        if (row_index == 2) or (row_index > 2):
            for column_index in range(1, 9+1):
                sheet.cell(row=row_index, column=column_index) \
                        .border = BORDER_NORMAL
    global KAMOKU_W, HOJO_W, KINGAKU_W
    global SUITOU_W, KUBUN_W
    global HIZUKE_W, TEKIYOU_W
    # 各列の横幅の設定
    sheet.column_dimensions['A'].width = HIZUKE_W
    sheet.column_dimensions['B'].width = KAMOKU_W
    sheet.column_dimensions['C'].width = HOJO_W
    sheet.column_dimensions['D'].width = KINGAKU_W
    sheet.column_dimensions['E'].width = KAMOKU_W
    sheet.column_dimensions['F'].width = HOJO_W
    sheet.column_dimensions['G'].width = KINGAKU_W
    sheet.column_dimensions['H'].width = TEKIYOU1_W
    sheet.column_dimensions['I'].width = TEKIYOU2_W
    # 印刷設定
    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    sheet.add_data_validation(dv_kamoku)
    sheet.add_data_validation(dv_hojo)

    return

def create_yokuki_fukugoushiwake_sheet(wb, sheet_name,
            dv_kamoku, dv_hojo):
    '''
    翌期入力用のExcelファイルに複合仕訳シートを作成する。

    Parameters
    ----------
    wb : workbook
        翌期入力用Excelファイル
    sheet_name : sheet_name
        複合仕訳シートのシート名
    dv_kamoku :
        勘定科目のドロップダウンリスト設定用データ
    dv_hojo :
        補助科目のドロップダウンリスト設定用データ

    Returns
    -------
    '''
    sheet = wb.create_sheet(title=sheet_name)
    sheet["A1"] = sheet_name

    sheet["A2"] = HIZUKE
    sheet["B2"] = KARIKATA_KAMOKU
    sheet["C2"] = KARIKATA_HOJO_KAMOKU
    sheet["D2"] = KARIKATA_KINGAKU
    sheet["E2"] = KASHIKATA_KAMOKU
    sheet["F2"] = KASHIKATA_HOJO_KAMOKU
    sheet["G2"] = KASHIKATA_KINGAKU
    sheet["H2"] = TEKIYOU
    sheet["I2"] = TEKIYOU2

    global TAKASA
    for row_index in range(50):
        sheet.row_dimensions[row_index + 1].height = TAKASA
        if row_index < 1:
            continue
        if row_index == 2:
            for column_index in range(1, 9+1):
                sheet.cell(row=row_index, column=column_index) \
                        .alignment = \
                        Alignment(horizontal="center")
        if row_index > 2:
            sheet.append(('', '', '', '', '', '', '', '', ''))
            # 書式の設定
            sheet.cell(row=row_index, column=1).number_format = FORMAT_HIZUKE
            sheet.cell(row=row_index, column=4).number_format = "#,##0"
            sheet.cell(row=row_index, column=7).number_format = "#,##0"
            # ドロップダウンリストの設定
            dv_kamoku.add(sheet.cell(row_index, 2))
            dv_hojo.add(sheet.cell(row_index, 3))
            dv_kamoku.add(sheet.cell(row_index, 5))
            dv_hojo.add(sheet.cell(row_index, 6))
        if (row_index == 2) or (row_index > 2):
            for column_index in range(1, 9+1):
                sheet.cell(row=row_index, column=column_index) \
                        .border = BORDER_NORMAL
    global KAMOKU_W, HOJO_W, KINGAKU_W
    global SUITOU_W, KUBUN_W
    global HIZUKE_W, TEKIYOU_W
    # 各列の横幅の設定
    sheet.column_dimensions['A'].width = HIZUKE_W
    sheet.column_dimensions['B'].width = KAMOKU_W
    sheet.column_dimensions['C'].width = HOJO_W
    sheet.column_dimensions['D'].width = KINGAKU_W
    sheet.column_dimensions['E'].width = KAMOKU_W
    sheet.column_dimensions['F'].width = HOJO_W
    sheet.column_dimensions['G'].width = KINGAKU_W
    sheet.column_dimensions['H'].width = TEKIYOU1_W
    sheet.column_dimensions['I'].width = TEKIYOU2_W
    # 印刷設定
    sheet.page_setup.orientation \
            = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    sheet.add_data_validation(dv_kamoku)
    sheet.add_data_validation(dv_hojo)

    return


def read_kihon(excel_file_name, sheet_name):
    '''
    Excelファイル内の基本シートを読込み、
    基本のデータを作成する。

    Parameters
    ----------
    excel_file_name : str
        読込むExcelファイル名。
    sheet_name : str
        基本のデータが入力されているシート名。

    Returns
    dantai_mei : str
        団体名。
    kishu_bi : str
        期首日。
    kimatsu_bi : str
        期末日
    kamoku_list: list of tuple of str, int, str, boolean
        勘定科目名、期首残高、貸借区分、貸借のタプルのリスト
    hojo_kamoku_list: list of tuple of str, str, int, str, boolean
        勘定科目名、補助科目名、期首残高、貸借区分、貸借のタプルのリスト
    suitou_list : list of tuple of str, str
        出納帳がある勘定科目、補助科目のタプルのリスト
    -------
    '''
    d.dprint_method_start()

    if not os.path.isfile(excel_file_name):
        e.eprint('ファイルがありません', excel_file_name)
        exit()
    openpyxl.reader.excel.warnings.simplefilter('ignore') # warning 対策　入力規則無視
    book = xl.load_workbook(excel_file_name, data_only=True)
    if sheet_name not in book.sheetnames:
        msg = "ファイル「{}」にシート「{}」が必要です。" \
                .format(excel_file_name, sheet_name)
        e.eprint('シートがありません', msg)
        exit()
    sheet = book[sheet_name]

    dantai_mei = sheet.cell(row=DANTAI_MEI_ROW, column=2).value
    kishu_bi = sheet.cell(row=KISHU_BI_ROW, column=2).value
    if isinstance(kishu_bi, str):
        try:
            kishu_bi = datetime.strptime(kishu_bi, '%Y-%m-%d')
        except:
            msg = "シート{}の{}行目{}列目の{}は期首日を示す文字列である必要があります。" \
                    .format(sheet_name, KISHU_BI_ROW, 2, kishu_bi)
            e.eprint('データが間違っています', msg)
            exit()
    kimatsu_bi = sheet.cell(row=KIMATSU_BI_ROW, column=2).value
    if isinstance(kimatsu_bi, str):
        try:
            kimatsu_bi = datetime.strptime(kimatsu_bi, '%Y-%m-%d')
        except:
            msg = "シート{}の{}行目{}列目の{}は期末日を示す文字列である必要があります。" \
                    .format(sheet_name, KIMATSU_BI_ROW, 2, kimatsu_bi)
            e.eprint('データが間違っています', msg)
            exit()
    if kishu_bi > kimatsu_bi:
        msg = "期首日{}より期末日{}が前の日付になっています。" \
                .format(kishu_bi, kimatsu_bi)
        e.eprint('データが間違っています', msg)
        exit()
        
    kamoku_list = []
    hojo_kamoku_list = []
    suitou_list = []
    hojo_goukei = 0
    for index in range(KAMOKU_TITLE_ROW + 1, sheet.max_row + 1):
        kamoku = sheet.cell(row=index, column=KAMOKU_COLUMN).value
        if (kamoku[0] == '+') or (kamoku[0] == '*'):
            continue

        hojo_kamoku = sheet.cell(row=index, \
                column=HOJO_KAMOKU_COLUMN).value
        taishaku_kubun = sheet.cell(row=index, \
                column=TAISHAKU_KUBUN_COLUMN).value
        taishaku_flag = henkan_taishaku_kubun(taishaku_kubun)
        suitou = sheet.cell(row=index, \
                column=SUITOU_COLUMN).value

        if (hojo_kamoku == None) \
                or (hojo_kamoku == KAMOKU_HOJO_GOUKEI):
            # 勘定科目の処理
            kamoku_zandaka_ = read_cell(excel_file_name, sheet_name,
                    sheet, index, KAMOKU_ZANDAKA_COLUMN)
            try:
                kamoku_zandaka = int(kamoku_zandaka_)
            except:
                msg = "{}行目{}列目のデータ{}は、整数である必要があります。" \
                        .format(index, KAMOKU_ZANDAKA_COLUMN,
                        kamoku_zandaka_)
                e.eprint('データが間違っています', msg)
                exit()
            kamoku_list.append( \
                    (kamoku, kamoku_zandaka, taishaku_kubun, taishaku_flag))
            if suitou == SUITOU_ARI:
                suitou_list.append((kamoku,))
            if hojo_kamoku == KAMOKU_HOJO_GOUKEI:
                if kamoku_zandaka != hojo_goukei:
                    msg = "科目の期首残高{:,d}と補助科目の期首残高の合計{:,d}が異なります。" \
                            .format(kamoku_zandaka, hojo_goukei)
                    e.eprint('データが間違っています', msg)
                    exit()
            hojo_goukei = 0
        else:
            # 補助科目の処理
            hojo_kamoku_zandaka_ = read_cell(excel_file_name, sheet_name,
                    sheet, index, HOJO_KAMOKU_ZANDAKA_COLUMN)
            try:
                hojo_kamoku_zandaka = int(hojo_kamoku_zandaka_)
            except:
                msg = "{}行目{}列目のデータ{}は、整数である必要があります。" \
                        .format(index, HOJO_KAMOKU_ZANDAKA_COLUMN,
                        hojo_kamoku_zandaka_)
                e.eprint('データが間違っています', msg)
                exit()
            hojo_kamoku_list.append( \
                    (kamoku, hojo_kamoku, \
                    hojo_kamoku_zandaka, taishaku_kubun, taishaku_flag))
            if suitou == SUITOU_ARI:
                suitou_list.append((kamoku, hojo_kamoku))
            hojo_goukei = hojo_goukei + hojo_kamoku_zandaka
    # d.dprint(dantai_mei)
    # d.dprint(kishu_bi)
    # d.dprint(kimatsu_bi)
    # d.dprint(kamoku_list)
    # d.dprint(hojo_kamoku_list)
    # d.dprint(suitou_list)
    
    # TODO 資産計などのチェック
    
    d.dprint_method_end()
    return dantai_mei, kishu_bi, kimatsu_bi, \
            kamoku_list, hojo_kamoku_list, suitou_list


def read_settei(excel_file_name, sheet_name):
    '''
    Excelの設定ファイル内の設定シートから、
    設定データ（ファイル名、セルの横幅、フォーマット）を読込む。

    Parameters
    ----------
    excel_file_name : str
        読込むExcelファイル名。
    sheet_name : str
        設定データが入力されているシート名。

    Returns
    -------
    '''
    d.dprint_method_start()

    if not os.path.isfile(excel_file_name):
        e.eprint('ファイルがありません', excel_file_name)
        exit()
    openpyxl.reader.excel.warnings.simplefilter('ignore') # warning 対策　入力規則無視
    book = xl.load_workbook(excel_file_name, data_only=True)
    if sheet_name not in book.sheetnames:
        msg = "ファイル「{}」にシート「{}」が必要です。" \
                .format(excel_file_name, sheet_name)
        e.eprint('シートがありません', msg)
        exit()
    sheet = book[sheet_name]

    global INPUT_FILE_NAME, SHISANHYOU_FILE_NAME
    global SOUKANJOU_MOTOCHOU_FILE_NAME, HOJO_MOTOCHOU_FILE_NAME
    global SHIWAKECHOU_FILE_NAME
    global YOKUNENDO_FILE_NAME
    INPUT_FILE_NAME = sheet.cell(row=1, column=2).value
    if INPUT_FILE_NAME == "" or INPUT_FILE_NAME == None:
        msg = "B1のセルには、会計データの入力ファイル名を入れてください。"
        e.eprint('ファイル名の指定がありません', msg)
        exit()
    SHISANHYOU_FILE_NAME = sheet.cell(row=2, column=2).value
    if SHISANHYOU_FILE_NAME == "" or SHISANHYOU_FILE_NAME == None:
        msg = "B2のセルには、残高試算表ファイル名を入れてください。"
        e.eprint('ファイル名の指定がありません', msg)
        exit()
    SOUKANJOU_MOTOCHOU_FILE_NAME = sheet.cell(row=3, column=2).value
    if SOUKANJOU_MOTOCHOU_FILE_NAME == "" or SOUKANJOU_MOTOCHOU_FILE_NAME == None:
        msg = "B3のセルには、総勘定元帳ファイル名を入れてください。"
        e.eprint('ファイル名の指定がありません', msg)
        exit()
    HOJO_MOTOCHOU_FILE_NAME = sheet.cell(row=4, column=2).value
    if HOJO_MOTOCHOU_FILE_NAME == "" or HOJO_MOTOCHOU_FILE_NAME == None:
        msg = "B4のセルには、補助元帳ファイル名を入れてください。"
        e.eprint('ファイル名の指定がありません', msg)
        exit()
    SHIWAKECHOU_FILE_NAME = sheet.cell(row=5, column=2).value
    if SHIWAKECHOU_FILE_NAME == "" or SHIWAKECHOU_FILE_NAME == None:
        msg = "B5のセルには、仕訳帳ファイル名を入れてください。"
        e.eprint('ファイル名の指定がありません', msg)
        exit()
    YOKUNENDO_FILE_NAME = sheet.cell(row=6, column=2).value
    if YOKUNENDO_FILE_NAME == "" or YOKUNENDO_FILE_NAME == None:
        msg = "B6のセルには、翌年度会計データのファイル名を入れてください。"
        e.eprint('ファイル名の指定がありません', msg)
        exit()

    global TAKASA
    TAKASA = read_cell(excel_file_name, sheet_name,
            sheet, 7, 2)

    global HIZUKE_W, BANGOU_W
    global KAMOKU_W, HOJO_W, KINGAKU_W
    global TEKIYOU1_W, TEKIYOU2_W, TEKIYOU_W
    global SUITOU_W, KUBUN_W
    global FORMAT_HIZUKE, FORMAT_KIKAN
    y_row = 8
    HIZUKE_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 1, 2)
    KAMOKU_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 2, 2)
    HOJO_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 3, 2)
    KINGAKU_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 4, 2)
    TEKIYOU1_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 5, 2)
    TEKIYOU2_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 6, 2)
    TEKIYOU_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 7, 2)
    SUITOU_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 8, 2)
    KUBUN_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 9, 2)
    BANGOU_W = read_cell(excel_file_name, sheet_name,
            sheet, y_row + 10, 2)

    FORMAT_HIZUKE = sheet.cell(row=19, column=2).value
    FORMAT_KIKAN = sheet.cell(row=20, column=2).value

    return

def read_cell(file_name, sheet_name, sheet, row, column):
    data = sheet.cell(row=row, column=column).value
    try:
        data_f = float(data)
    except ValueError:
        cell_name = chr(ord('A') + column - 1) + str(row)
        msg = "ファイル「{}」のシート「{}」のセル「{}」" \
                .format(file_name, sheet_name, cell_name)
        e.eprint("数字である必要があります", msg)
        exit()
    return data_f


if __name__ == '__main__':
    read_settei(SETTEI_FILE_NAME, SETTEI_SHEET_NAME)
    dantai_mei, kishu_bi, kimatsu_bi, \
            kamoku_list, hojo_kamoku_list, suitou_list \
            = read_kihon(INPUT_FILE_NAME, KIHON_SHEET_NAME)
    d.dprint_name("kishu_bi", kishu_bi)
    d.dprint_name("kimatsu_bi", kimatsu_bi)
    d.dprint(type(kishu_bi))

    suitou_chou_list = []
    suitou_kamoku_list = []
    for suitou in suitou_list:
        if len(suitou) == 1:
            suitou_chou = read_suitou(INPUT_FILE_NAME,
                    suitou[0], suitou[0], '')
            suitou_kamoku_list.append((suitou[0], ''))
        else:
            suitou_chou = read_suitou(INPUT_FILE_NAME,
                    suitou[0] + '_' + suitou[1],
                    suitou[0], suitou[1])
            suitou_kamoku_list.append((suitou[0], suitou[1]))
        suitou_chou_list.append(suitou_chou)

    # 単一仕訳データを読込む
    tanitsu = read_tanitsu_shiwake(INPUT_FILE_NAME,
            TANITSU_SHEET_NAME)
    suitou_chou_list.append(tanitsu)

    # 複合仕訳データを読込む
    fukugou = read_fukugou_shiwake(INPUT_FILE_NAME,
            FUKUGOU_SHEET_NAME)

    shiwake_chou = ketsugou_shiwake(suitou_chou_list,
            suitou_kamoku_list, fukugou)
    for suitou_chou in suitou_chou_list:
        del suitou_chou
    d.dprint(shiwake_chou)

    # 総勘定元帳、補助元帳データ作成
    d.dprint_name("kishu_bi", kishu_bi)
    d.dprint_name("kimatsu_bi", kimatsu_bi)
    d.dprint(type(kishu_bi))
    soukanjou_motochou_list = []
    shisanhyou_list = []
#         kamoku_list: list of tuple of str, int, str, boolean
#         勘定科目名、期首残高、貸借区分、貸借のタプルのリスト
    taishaku_kubun = kamoku_list[0][2]
    kishu_kubun_goukei = 0
    karikata_kubun_goukei = 0
    kashikata_kubun_goukei = 0
    kimatsu_kubun_goukei = 0
    kishu_junshisan = 0
    karikata_junshisan = 0
    kashikata_junshisan = 0
    kimatsu_junshisan = 0

    for kamoku in kamoku_list:
        soukanjou_motochou, zandaka, karikata_goukei, kashikata_goukei \
                = sakusei_soukanjou_motochou(shiwake_chou, kamoku,
                kishu_bi, kimatsu_bi)
        soukanjou_motochou_list.append(
                (kamoku, soukanjou_motochou))
        if taishaku_kubun != kamoku[2]:
            shisanhyou_list.append((
                    "++ "+taishaku_kubun+"計 ++",
                    kishu_kubun_goukei, karikata_kubun_goukei,
                    kashikata_kubun_goukei, kimatsu_kubun_goukei))
            if taishaku_kubun == TAISHAKU_KUBUN_SHISAN:
                kishu_junshisan = kishu_kubun_goukei
                karikata_junshisan = karikata_kubun_goukei
                kashikata_junshisan = kashikata_kubun_goukei
                kimatsu_junshisan = kimatsu_kubun_goukei
            if taishaku_kubun == TAISHAKU_KUBUN_SHUNYU:
                kishu_shushi = kishu_kubun_goukei
                karikata_shushi = karikata_kubun_goukei
                kashikata_shushi = kashikata_kubun_goukei
                kimatsu_shushi = kimatsu_kubun_goukei
            if taishaku_kubun == TAISHAKU_KUBUN_FUSAI:
                kishu_junshisan -= kishu_kubun_goukei
                karikata_junshisan -= karikata_kubun_goukei
                kashikata_junshisan -= kashikata_kubun_goukei
                kimatsu_junshisan -= kimatsu_kubun_goukei
                shisanhyou_list.append((
                        "** "+TAISHAKU_KUBUN_JUNSHISAN+" **",
                        kishu_junshisan, karikata_junshisan,
                        kashikata_junshisan, kimatsu_junshisan))
            kishu_kubun_goukei = 0
            karikata_kubun_goukei = 0
            kashikata_kubun_goukei = 0
            kimatsu_kubun_goukei = 0
            taishaku_kubun = kamoku[2]
        shisanhyou_list.append((kamoku[0], kamoku[1],
                karikata_goukei, kashikata_goukei, zandaka))
        kishu_kubun_goukei += kamoku[1]
        karikata_kubun_goukei += karikata_goukei
        kashikata_kubun_goukei += kashikata_goukei
        kimatsu_kubun_goukei += zandaka
    shisanhyou_list.append((
            "++ "+taishaku_kubun+"計 ++",
            kishu_kubun_goukei, karikata_kubun_goukei,
            kashikata_kubun_goukei, kimatsu_kubun_goukei))
    kishu_shushi -= kishu_kubun_goukei
    karikata_shushi -= karikata_kubun_goukei
    kashikata_shushi -= kashikata_kubun_goukei
    kimatsu_shushi -= kimatsu_kubun_goukei
    shisanhyou_list.append((
            "** "+TAISHAKU_KUBUN_SHUSHI+" **",
            kishu_shushi, karikata_shushi,
            kashikata_shushi, kimatsu_shushi))

    hojo_motochou_list = []
    hojo_ichiran_list = []
    kamoku_mei = hojo_kamoku_list[0][0]
    kishu_kamoku_goukei = 0
    karikata_kamoku_goukei = 0
    kashikata_kamoku_goukei = 0
    kimatsu_kamoku_goukei = 0
    for hojo_kamoku in hojo_kamoku_list:
        hojo_motochou, zandaka, karikata_goukei, kashikata_goukei \
                = sakusei_hojo_motochou(shiwake_chou, hojo_kamoku,
                kishu_bi, kimatsu_bi)
        hojo_motochou_list.append((hojo_kamoku, hojo_motochou))
        if kamoku_mei != hojo_kamoku[0]:
            hojo_ichiran_list.append(
                    (kamoku_mei, "【合計】", \
                    kishu_kamoku_goukei,
                    karikata_kamoku_goukei,
                    kashikata_kamoku_goukei,
                    kimatsu_kamoku_goukei))
            kamoku_mei = hojo_kamoku[0]
            kishu_kamoku_goukei = 0
            karikata_kamoku_goukei = 0
            kashikata_kamoku_goukei = 0
            kimatsu_kamoku_goukei = 0
        hojo_ichiran_list.append((hojo_kamoku[0], hojo_kamoku[1], \
                hojo_kamoku[2], karikata_goukei,
                kashikata_goukei, zandaka))
        kishu_kamoku_goukei += hojo_kamoku[2]
        karikata_kamoku_goukei += karikata_goukei
        kashikata_kamoku_goukei += kashikata_goukei
        kimatsu_kamoku_goukei += zandaka
    hojo_ichiran_list.append(
            (kamoku_mei, "【合計】", \
            kishu_kamoku_goukei,
            karikata_kamoku_goukei,
            kashikata_kamoku_goukei,
            kimatsu_kamoku_goukei))
    kishu_kamoku_goukei = 0
    karikata_kamoku_goukei = 0
    kashikata_kamoku_goukei = 0
    kimatsu_kamoku_goukei = 0

    # Excelに出力
    d.dprint_name("kishu_bi", kishu_bi)
    d.dprint_name("kimatsu_bi", kimatsu_bi)
    d.dprint(type(kishu_bi))
    if SHISANHYOU_FILE_NAME != None:
        save_shisanhyou_file(SHISANHYOU_FILE_NAME,
                dantai_mei, kishu_bi, kimatsu_bi,
                shisanhyou_list, hojo_ichiran_list)
    if SOUKANJOU_MOTOCHOU_FILE_NAME != None:
        save_soukanjou_motochou_file(SOUKANJOU_MOTOCHOU_FILE_NAME,
                dantai_mei, kishu_bi, kimatsu_bi,
                soukanjou_motochou_list)
    if HOJO_MOTOCHOU_FILE_NAME != None:
        save_hojo_motochou_file(HOJO_MOTOCHOU_FILE_NAME,
                dantai_mei, kishu_bi, kimatsu_bi,
                hojo_motochou_list)
    if SHIWAKECHOU_FILE_NAME != None:
        save_shiwakechou_file(SHIWAKECHOU_FILE_NAME,
                dantai_mei, kishu_bi, kimatsu_bi,
                shiwake_chou)

    # 翌期用Excel作成
    if YOKUNENDO_FILE_NAME != None:
        save_yokuki_kihon(YOKUNENDO_FILE_NAME,
                dantai_mei, kishu_bi, kimatsu_bi,
                suitou_list, shisanhyou_list, hojo_ichiran_list,
                kamoku_list, hojo_kamoku_list)

